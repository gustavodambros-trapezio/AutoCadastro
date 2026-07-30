# -*- coding: utf-8 -*-
"""
Site local do AutoCadastro Protheus.

Roda no PC da sala de servidores, aberto para a rede:
    python app_autocadastro.py       ->  http://IP_DO_SERVIDOR:8025

Telas:
  - Novo cadastro: escolhe a FILIAL (lista real do Protheus) e preenche a
    grade NOME | CPF | FUNÇÃO (colar do Excel espalha pelas células).
    Ao enviar, roda o robô e mostra o resultado ao vivo.
  - Histórico: todas as execuções, com todas as colunas.
  - Usuários criados pela automação: SÓ os usuários com STATUS=CRIADO, com
    abas por filial + Exportar Excel (dentro da página). Cadastros que não
    concluíram aparecem apenas no Histórico de execuções.

IMPORTANTE — o robô usa a sessão do Protheus JÁ ABERTA E LOGADA no Chrome do
servidor (nenhuma senha do Protheus fica guardada). O site mostra esse aviso
em todas as telas e verifica a conexão antes de criar.

Requisitos: pip install flask waitress openpyxl selenium
"""

import base64
import io
import json
import os
import re
import socket
import sqlite3
import subprocess
import sys
import threading
from datetime import datetime
from functools import wraps

from flask import (Flask, Response, redirect, render_template_string, request,
                   send_file, url_for)

# este arquivo mora em site/; dados e configs ficam na RAIZ do projeto
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "autocadastro.db")
SCRIPT = os.path.join(BASE_DIR, "robo", "protheus_criar_usuario.py")
FILIAIS_PATH = os.path.join(BASE_DIR, "filiais.json")

try:
    with open(os.path.join(BASE_DIR, "protheus_config.json"), encoding="utf-8") as _f:
        _cfg = json.load(_f)
except (OSError, ValueError):
    _cfg = {}

PORTA = int(os.environ.get("AUTOCADASTRO_PORTA") or _cfg.get("porta_site") or 8025)
SITE_USUARIO = os.environ.get("AUTOCADASTRO_USER") or _cfg.get("site_usuario") or "admin"
SITE_SENHA = os.environ.get("AUTOCADASTRO_PASS") or _cfg.get("site_senha") or ""
CHROME_DEBUG = _cfg.get("chrome_debug", "127.0.0.1:9222")
TIMEOUT_EXECUCAO_S = 3600

app = Flask(__name__)
_trava_execucao = threading.Lock()
# processo do robô em andamento (para o botão "Parar execução")
_proc_atual = {"execucao_id": None, "proc": None}
_paradas = set()   # execuções paradas pelo botão (para dar o status certo)


# ----------------------------------------------------------------------------
# LOGIN DO SITE (básico — só para não ficar aberto a qualquer um da rede)
# ----------------------------------------------------------------------------
def exige_login(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if not SITE_SENHA:
            return f(*a, **kw)
        auth = request.authorization
        if auth and auth.username == SITE_USUARIO and auth.password == SITE_SENHA:
            return f(*a, **kw)
        return Response(
            "Acesso restrito ao AutoCadastro.", 401,
            {"WWW-Authenticate": 'Basic realm="AutoCadastro Protheus"'})
    return wrapper


# ----------------------------------------------------------------------------
# FILIAIS (lista real capturada do Protheus)
# ----------------------------------------------------------------------------
def carregar_filiais():
    try:
        with open(FILIAIS_PATH, encoding="utf-8") as f:
            filiais = json.load(f)
    except (OSError, ValueError):
        return []
    return [{"codigo": x.get("codigo", ""), "descricao": x.get("descricao", "")}
            for x in filiais if x.get("codigo")]


def nome_da_filial(codigo):
    for f in carregar_filiais():
        if f["codigo"] == codigo:
            return f["descricao"]
    return ""


def aprender_filial(codigo, descricao):
    """
    Guarda no filiais.json o código + descrição que o robô leu do Protheus.
    Assim a lista de sugestões se completa sozinha conforme filiais novas vão
    sendo usadas — não é preciso manter a lista à mão quando abrir posto novo.
    """
    codigo = (codigo or "").strip().upper()
    descricao = (descricao or "").strip()
    if not codigo or not descricao:
        return False
    filiais = carregar_filiais()
    for f in filiais:
        if f["codigo"] == codigo:
            if f["descricao"] == descricao:
                return False
            f["descricao"] = descricao
            break
    else:
        filiais.append({"codigo": codigo, "descricao": descricao})
    filiais.sort(key=lambda f: f["codigo"])
    try:
        with open(FILIAIS_PATH, "w", encoding="utf-8") as fh:
            json.dump(filiais, fh, ensure_ascii=False, indent=2)
        return True
    except OSError:
        return False


# ----------------------------------------------------------------------------
# BANCO
# ----------------------------------------------------------------------------
def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def init_db():
    with db() as con:
        con.executescript("""
        CREATE TABLE IF NOT EXISTS execucoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            iniciada TEXT NOT NULL,
            finalizada TEXT,
            filial TEXT NOT NULL,
            filial_nome TEXT DEFAULT '',
            total INTEGER DEFAULT 0,
            criados INTEGER DEFAULT 0,
            erros INTEGER DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'RODANDO'
        );
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            execucao_id INTEGER NOT NULL REFERENCES execucoes(id),
            filial TEXT NOT NULL,
            filial_nome TEXT DEFAULT '',
            nome TEXT NOT NULL,
            cpf TEXT NOT NULL,
            funcao TEXT NOT NULL,
            usuario TEXT DEFAULT '',
            senha TEXT DEFAULT '',
            grupo_codigo TEXT DEFAULT '',
            grupo_nome TEXT DEFAULT '',
            prioriza_grupo TEXT DEFAULT '',
            forcar_troca_senha TEXT DEFAULT '',
            codigo_banco TEXT DEFAULT '',
            id_usuario TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'PROCESSANDO',
            criado_em TEXT
        );
        """)
        # migração leve para bancos criados antes das colunas de nome da filial
        for tabela in ("execucoes", "usuarios"):
            cols = {r["name"] for r in con.execute(f"PRAGMA table_info({tabela})")}
            if "filial_nome" not in cols:
                con.execute(f"ALTER TABLE {tabela} ADD COLUMN filial_nome TEXT DEFAULT ''")


def agora():
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def so_digitos(texto):
    return re.sub(r"\D", "", texto or "")


# ----------------------------------------------------------------------------
# ESTADO DO ROBÔ / PROTHEUS
# ----------------------------------------------------------------------------
def chrome_conectado():
    """O Chrome logado no Protheus está acessível na porta de depuração?"""
    try:
        host, porta = CHROME_DEBUG.split(":")
        with socket.create_connection((host, int(porta)), timeout=2):
            return True
    except Exception:
        return False


# ----------------------------------------------------------------------------
# EXECUÇÃO
# ----------------------------------------------------------------------------
def parse_linhas(texto):
    """Cada linha: NOME;CPF;FUNÇÃO (aceita TAB — colar do Excel — ou ';')."""
    linhas, problemas = [], []
    for n, bruta in enumerate((texto or "").splitlines(), start=1):
        bruta = bruta.strip()
        if not bruta:
            continue
        partes = [p.strip() for p in re.split(r"[\t;]", bruta) if p.strip()]
        if len(partes) < 3:
            problemas.append(f"Linha {n}: preciso de NOME;CPF;FUNÇÃO — recebi: {bruta!r}")
            continue
        nome, cpf, funcao = partes[0], partes[1], " ".join(partes[2:])
        if not so_digitos(cpf):
            problemas.append(f"Linha {n}: CPF sem números: {bruta!r}")
            continue
        linhas.append({"nome": nome.upper(), "cpf": cpf, "funcao": funcao.upper()})
    return linhas, problemas


def _grava_resultado(con, r):
    """Grava o resultado de um usuário (parcial ou final) na linha dele."""
    con.execute("""
        UPDATE usuarios SET usuario=?, senha=?, grupo_codigo=?, grupo_nome=?,
            prioriza_grupo=?, forcar_troca_senha=?, codigo_banco=?,
            id_usuario=?, status=?, criado_em=?
        WHERE id=?""",
        (r.get("usuario", ""), r.get("senha", ""), r.get("grupo_codigo", ""),
         r.get("grupo_nome", ""), r.get("prioriza_grupo", ""),
         r.get("forcar_troca_senha", ""), r.get("codigo_banco", ""),
         r.get("id_usuario", ""), r.get("status", "ERRO: sem status"),
         agora(), r.get("row_number")))


def rodar_execucao(execucao_id, pendentes):
    try:
        lote = [{"row_number": u["id"], "nome": u["nome"], "cpf": u["cpf"],
                 "funcao": u["funcao"], "filial": u["filial"]} for u in pendentes]
        b64 = base64.b64encode(json.dumps(lote).encode("utf-8")).decode("ascii")
        # Popen (e não run): o botão "Parar execução" precisa poder matar o
        # robô, e o stdout é lido LINHA A LINHA para atualizar cada usuário na
        # hora ("@@PARCIAL@@{json}" por usuário; a última linha é o JSON final).
        proc = subprocess.Popen(
            [sys.executable, SCRIPT, "--json-b64", b64],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            text=True, encoding="utf-8", errors="replace", cwd=BASE_DIR)
        _proc_atual.update(execucao_id=execucao_id, proc=proc)
        # o stderr precisa ser drenado em paralelo (se o cano encher, o robô
        # trava no write); guardamos as linhas para diagnosticar erros
        stderr_linhas = []
        drenador = threading.Thread(
            target=lambda: stderr_linhas.extend(proc.stderr), daemon=True)
        drenador.start()
        # rede de segurança para o robô pendurado (o timeout do communicate
        # não existe mais no modo linha a linha)
        matador = threading.Timer(TIMEOUT_EXECUCAO_S, proc.kill)
        matador.start()
        saida_final = ""
        try:
            for linha in proc.stdout:
                linha = linha.strip()
                if linha.startswith("@@PARCIAL@@"):
                    try:
                        r = json.loads(linha[len("@@PARCIAL@@"):])
                    except ValueError:
                        continue
                    with db() as con:
                        _grava_resultado(con, r)
                elif linha.startswith("[") and linha.endswith("]"):
                    saida_final = linha
            proc.wait()
            drenador.join(timeout=5)
        finally:
            matador.cancel()
            _proc_atual.update(execucao_id=None, proc=None)
        if execucao_id in _paradas:
            raise RuntimeError("cancelada pelo usuário no site")

        if not saida_final:
            raise RuntimeError(
                f"O robô não devolveu resultado. Detalhe: {''.join(stderr_linhas)[-600:]}")
        resultados = json.loads(saida_final)

        # o robô lê a descrição da filial no próprio Protheus: aproveita para
        # completar a lista de sugestões (filiais novas entram sozinhas)
        for r in resultados:
            if r.get("filial") and r.get("filial_nome"):
                if aprender_filial(r["filial"], r["filial_nome"]):
                    with db() as con:
                        con.execute(
                            "UPDATE execucoes SET filial_nome=? WHERE filial=? AND filial_nome=''",
                            (r["filial_nome"], r["filial"]))
                break

        with db() as con:
            for r in resultados:
                _grava_resultado(con, r)
            _finalizar(con, execucao_id, "CONCLUIDA")
    except Exception as e:
        parada = execucao_id in _paradas
        msg = ("ERRO: execução cancelada pelo usuário no site" if parada
               else f"ERRO: {e}")
        with db() as con:
            con.execute(
                "UPDATE usuarios SET status=?, criado_em=? WHERE execucao_id=? AND status='PROCESSANDO'",
                (msg, agora(), execucao_id))
            _finalizar(con, execucao_id, "CANCELADO" if parada else "FALHOU")
    finally:
        _paradas.discard(execucao_id)
        _trava_execucao.release()


def _finalizar(con, execucao_id, status):
    q = lambda sql: con.execute(sql, (execucao_id,)).fetchone()["c"]  # noqa: E731
    tot = q("SELECT COUNT(*) c FROM usuarios WHERE execucao_id=?")
    ok = q("SELECT COUNT(*) c FROM usuarios WHERE execucao_id=? AND status='CRIADO'")
    err = q("SELECT COUNT(*) c FROM usuarios WHERE execucao_id=? AND status LIKE 'ERRO%'")
    # pedido do usuário (30/07/2026): bastou 1 erro, a execução é FALHOU
    if status == "CONCLUIDA" and err:
        status = "FALHOU"
    con.execute("UPDATE execucoes SET finalizada=?, total=?, criados=?, erros=?, status=? WHERE id=?",
                (agora(), tot, ok, err, status, execucao_id))


# ----------------------------------------------------------------------------
# TEMPLATES
# ----------------------------------------------------------------------------
ESTILO = """
<style>
  * { box-sizing: border-box; }
  body { font-family: Arial, Helvetica, sans-serif; margin: 0; background: #f4f6f8; color: #222; }
  header { background: #0c3c60; color: #fff; padding: 14px 24px; display: flex; align-items: center; gap: 18px; flex-wrap: wrap; }
  header h1 { font-size: 18px; margin: 0 18px 0 0; }
  header a { color: #cfe3f5; text-decoration: none; font-size: 14px; padding: 6px 10px; border-radius: 4px; }
  header a:hover, header a.ativo { background: #14547f; color: #fff; }
  main { max-width: 1250px; margin: 22px auto; padding: 0 16px; }
  .cartao { background: #fff; border: 1px solid #dde3e8; border-radius: 8px; padding: 20px; margin-bottom: 20px; }
  label { display: block; font-weight: bold; margin: 12px 0 4px; font-size: 14px; }
  select, textarea, input[type=text] { width: 100%; padding: 8px; border: 1px solid #bbb; border-radius: 4px; font-size: 14px; }
  input[type=text] { font-family: Consolas, monospace; text-transform: uppercase; }
  textarea { min-height: 170px; font-family: Consolas, monospace; }
  button, .botao { background: #0c6b3d; color: #fff; border: 0; padding: 11px 24px; border-radius: 4px; font-size: 15px; cursor: pointer; text-decoration: none; display: inline-block; margin-top: 14px; }
  button:hover, .botao:hover { background: #0e7f49; }
  button[disabled] { background: #9bb0a5; cursor: not-allowed; }
  table { border-collapse: collapse; width: 100%; font-size: 13px; background: #fff; }
  th, td { border: 1px solid #d7dde2; padding: 6px 8px; text-align: left; white-space: nowrap; }
  th { background: #eef2f5; }
  .rolagem { overflow-x: auto; }
  .st-CRIADO { background: #c6efce; color: #006100; font-weight: bold; }
  .st-ERRO { background: #ffc7ce; color: #9c0006; }
  .st-PROCESSANDO { background: #fff2cc; color: #7f6000; }
  .st-AGUARDANDO { background: #ddebf7; color: #1f4e79; }
  .st-JAEXISTE { background: #e2e2e2; color: #444; }
  .aviso { background: #fff3cd; border: 1px solid #ffe08a; padding: 11px 15px; border-radius: 6px; margin-bottom: 16px; font-size: 14px; }
  .erro { background: #ffd8d8; border: 1px solid #f5a9a9; padding: 11px 15px; border-radius: 6px; margin-bottom: 16px; font-size: 14px; }
  .ok { background: #d9f2e3; border: 1px solid #a5d9bd; padding: 11px 15px; border-radius: 6px; margin-bottom: 16px; font-size: 14px; }
  .abas { margin-bottom: 14px; display: flex; flex-wrap: wrap; gap: 6px; }
  .abas a { background: #dfe7ee; color: #0c3c60; padding: 7px 14px; border-radius: 6px 6px 0 0; text-decoration: none; font-size: 13px; }
  .abas a.ativa { background: #0c3c60; color: #fff; font-weight: bold; }
  .mini { color: #667; font-size: 12px; line-height: 1.5; }
  code { background: #eef2f5; padding: 1px 5px; border-radius: 3px; font-size: 12px; }
  table.grade { table-layout: fixed; }
  table.grade th { text-align: center; }
  table.grade td { padding: 0; }
  table.grade td.num { padding: 6px 4px; background: #eef2f5; color: #667; text-align: center; font-size: 12px; }
  table.grade input { border: 0; border-radius: 0; width: 100%; padding: 7px 8px; font-family: Consolas, monospace; font-size: 13px; text-transform: uppercase; outline: none; background: transparent; }
  table.grade input:focus { background: #eaf3ff; }
  button.parar { background: #b33232; }
  button.parar:hover { background: #8f2020; }
  button.apagar { background: none; border: 0; color: #b33; cursor: pointer; margin: 0; padding: 5px 8px; font-size: 15px; line-height: 1; }
  button.apagar:hover { background: #ffd8d8; border-radius: 3px; }
  form.form-apagar { margin: 0; display: inline; }
</style>
"""

BASE = """
<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <title>AutoCadastro Protheus</title>
  {% if refresh %}<meta http-equiv="refresh" content="5">{% endif %}
  """ + ESTILO + """
</head>
<body>
<header>
  <h1>AutoCadastro Protheus</h1>
  <a href="{{ url_for('index') }}" class="{{ 'ativo' if aba=='novo' }}">+ Novo cadastro</a>
  <a href="{{ url_for('historico') }}" class="{{ 'ativo' if aba=='historico' }}">Histórico</a>
  <a href="{{ url_for('planilha') }}" class="{{ 'ativo' if aba=='planilha' }}">Usuários criados pela automação</a>
</header>
<main>
{% if not conectado %}
  <div class="aviso">ℹ️ O Chrome do robô está fechado — <b>sem problema</b>:
  ao criar, o robô abre o Chrome e loga sozinho com o usuário-robô
  (<code>AUTO.PROTHEUS</code>). A primeira execução só demora ~1 minuto a mais.</div>
{% else %}
  <div class="ok">✅ Sessão do Protheus conectada. O robô vai usar a janela do Chrome que está logada
  (e reloga sozinho se a sessão tiver caído).</div>
{% endif %}
{% if rodando %}
  <div class="aviso">⏳ Execução em andamento (nº {{ rodando['id'] }}, filial {{ rodando['filial'] }}).
  <a href="{{ url_for('ver_execucao', execucao_id=rodando['id']) }}">Acompanhar</a></div>
{% endif %}
{{ conteudo | safe }}
</main>
</body>
</html>
"""

TABELA_USUARIOS = """
<div class="rolagem"><table>
<tr><th>#</th><th>FILIAL</th><th>NOME</th><th>CPF</th><th>FUNÇÃO</th><th>USUARIO</th>
<th>SENHA</th><th>GRUPO</th><th>GRUPO NOME</th><th>PRIORIZA</th><th>TROCA SENHA</th>
<th>CÓD. BANCO</th><th>ID USUÁRIO</th><th>STATUS</th><th>QUANDO</th>
{% if apagar %}<th></th>{% endif %}</tr>
{% for u in usuarios %}
<tr>
  <td>{{ loop.index }}</td>
  <td title="{{ u['filial_nome'] or '' }}">{{ u['filial'] }}</td>
  <td>{{ u['nome'] }}</td><td>{{ u['cpf'] }}</td><td>{{ u['funcao'] }}</td>
  <td>{{ u['usuario'] }}</td><td>{{ u['senha'] }}</td><td>{{ u['grupo_codigo'] }}</td>
  <td>{{ u['grupo_nome'] }}</td><td>{{ u['prioriza_grupo'] }}</td>
  <td>{{ u['forcar_troca_senha'] }}</td><td>{{ u['codigo_banco'] }}</td>
  <td>{{ u['id_usuario'] }}</td>
  <td class="st-{{ classe_status(u['status']) }}">{{ u['status'] }}</td>
  <td>{{ u['criado_em'] or '' }}</td>
  {% if apagar %}
  <td>
    <form method="post" class="form-apagar"
          action="{{ url_for('apagar_usuario', usuario_id=u['id'], filial=request.args.get('filial', '')) }}"
          onsubmit="return confirm('Apagar {{ u['nome'] }} SOMENTE do banco do site?\\n\\nO usuário {{ u['usuario'] or '(sem login)' }} continua existindo no Protheus. O CPF fica liberado para um novo cadastro.')">
      <button type="submit" class="apagar" title="Apagar do banco do site (não mexe no Protheus)">&#10005;</button>
    </form>
  </td>
  {% endif %}
</tr>
{% endfor %}
</table></div>
"""


SCRIPT_GRADE = """
<script>
(function () {
  var corpo = document.getElementById('grade-corpo');
  var form = document.getElementById('form-criar');
  if (!corpo || !form) return;
  var COLS = 3;  // NOME, CPF, FUNÇÃO

  function inputs(tr) { return tr.querySelectorAll('input'); }

  function renumera() {
    for (var i = 0; i < corpo.rows.length; i++)
      corpo.rows[i].cells[0].textContent = i + 1;
  }

  function novaLinha() {
    var tr = corpo.insertRow(-1);
    tr.insertCell(-1).className = 'num';
    for (var c = 0; c < COLS; c++) {
      var inp = document.createElement('input');
      inp.type = 'text';
      inp.dataset.col = c;
      inp.autocomplete = 'off';
      tr.insertCell(-1).appendChild(inp);
    }
    var tdx = tr.insertCell(-1);
    tdx.className = 'num';
    var bx = document.createElement('button');
    bx.type = 'button';
    bx.className = 'apagar';
    bx.textContent = '\\u2715';
    bx.title = 'Remover linha';
    bx.onclick = function () {
      if (corpo.rows.length > 1) tr.remove();
      else inputs(tr).forEach(function (i) { i.value = ''; });
      renumera();
    };
    tdx.appendChild(bx);
    renumera();
    return tr;
  }

  // a última linha sempre fica em branco, pronta para a próxima pessoa
  function garanteVazia() {
    var ult = corpo.rows[corpo.rows.length - 1];
    var temAlgo = Array.prototype.some.call(inputs(ult),
      function (i) { return i.value.trim() !== ''; });
    if (temAlgo) novaLinha();
  }

  corpo.addEventListener('paste', function (e) {
    var alvo = e.target;
    if (alvo.tagName !== 'INPUT') return;
    var texto = (e.clipboardData || window.clipboardData).getData('text');
    if (!/[\\t\\r\\n;]/.test(texto)) return;  // valor simples: colagem normal
    e.preventDefault();
    var linhas = texto.replace(/\\r/g, '').split('\\n')
      .filter(function (l) { return l.trim() !== ''; });
    var tr0 = alvo.closest('tr');
    var lin0 = Array.prototype.indexOf.call(corpo.rows, tr0);
    var col0 = parseInt(alvo.dataset.col, 10);
    linhas.forEach(function (lt, i) {
      var tr = corpo.rows[lin0 + i] || novaLinha();
      var ins = inputs(tr);
      lt.split(/[\\t;]/).forEach(function (v, j) {
        var c = col0 + j;
        if (c < COLS) ins[c].value = v.trim();
      });
    });
    garanteVazia();
  });

  corpo.addEventListener('input', garanteVazia);

  corpo.addEventListener('keydown', function (e) {
    if (e.key !== 'Enter' || e.target.tagName !== 'INPUT') return;
    e.preventDefault();
    var tr = e.target.closest('tr');
    var lin = Array.prototype.indexOf.call(corpo.rows, tr);
    var prox = corpo.rows[lin + 1] || novaLinha();
    inputs(prox)[parseInt(e.target.dataset.col, 10)].focus();
  });

  form.addEventListener('submit', function (e) {
    var saida = [];
    var incompletas = [];
    Array.prototype.forEach.call(corpo.rows, function (tr, i) {
      var v = Array.prototype.map.call(inputs(tr),
        function (inp) { return inp.value.trim().toUpperCase(); });
      if (v.every(function (x) { return x === ''; })) return;
      if (v.some(function (x) { return x === ''; })) incompletas.push(i + 1);
      saida.push(v.join(';'));
    });
    if (incompletas.length) {
      e.preventDefault();
      alert('Preencha NOME, CPF e FUNÇÃO na(s) linha(s): ' + incompletas.join(', '));
      return;
    }
    if (!saida.length) {
      e.preventDefault();
      alert('Preencha ao menos uma linha.');
      return;
    }
    document.getElementById('linhas').value = saida.join('\\n');
  });

  for (var i = 0; i < 5; i++) novaLinha();
})();
</script>
"""


def classe_status(status):
    s = (status or "").upper()
    if s == "CRIADO":
        return "CRIADO"
    if s.startswith("ERRO"):
        return "ERRO"
    if s.startswith("AGUARDANDO"):
        return "AGUARDANDO"
    if s.startswith("JÁ EXISTE") or s.startswith("JA EXISTE"):
        return "JAEXISTE"
    return "PROCESSANDO"


def render(conteudo, aba="", refresh=False):
    with db() as con:
        rodando = con.execute(
            "SELECT * FROM execucoes WHERE status='RODANDO' ORDER BY id DESC LIMIT 1").fetchone()
    return render_template_string(BASE, conteudo=conteudo, aba=aba, refresh=refresh,
                                  rodando=rodando, conectado=chrome_conectado())


def render_tabela(usuarios, apagar=False):
    return render_template_string(TABELA_USUARIOS, usuarios=usuarios,
                                  classe_status=classe_status, apagar=apagar)


# ----------------------------------------------------------------------------
# ROTAS
# ----------------------------------------------------------------------------
@app.route("/")
@exige_login
def index():
    filiais = carregar_filiais()
    # Lista como SUGESTÃO, não como camisa de força: o levantamento de filiais
    # é parcial (a consulta do Protheus carrega por página), e o código é o
    # que o pessoal usa para nomear as páginas — então dá para digitar direto.
    opcoes = "".join(
        f'<option value="{f["codigo"]}">{f["codigo"]} — {f["descricao"]}</option>'
        for f in filiais)
    erro = request.args.get("erro", "")
    conteudo = f"""
    {f'<div class="erro">{erro}</div>' if erro else ''}
    <div class="cartao">
      <h2>Criar usuários no Protheus</h2>
      <form method="post" action="{url_for('criar')}" id="form-criar">
        <label>Filial — código do Protheus (ex.: 01DOMA0001)</label>
        <input type="text" name="filial" list="filiais" required
               autocomplete="off" placeholder="01DOMA0001"
               pattern="[0-9A-Za-z]{{6,20}}"
               title="Use o código da filial no Protheus, ex.: 01DOMA0001">
        <datalist id="filiais">{opcoes}</datalist>
        <p class="mini">Pode digitar qualquer código, mesmo que não esteja na
        lista de sugestões ({len(filiais)} conhecidas). Se o código não existir
        no Protheus, ninguém é criado e as linhas voltam como
        <b>ERRO: FILIAL NÃO EXISTE NO PROTHEUS</b>.</p>
        <label>Funcionários — preencha ou cole do Excel direto na tabela</label>
        <div class="rolagem">
          <table class="grade" id="grade">
            <thead><tr>
              <th style="width:38px">#</th><th>NOME</th>
              <th style="width:180px">CPF</th><th style="width:220px">FUNÇÃO</th>
              <th style="width:38px"></th>
            </tr></thead>
            <tbody id="grade-corpo"></tbody>
          </table>
        </div>
        <p class="mini">Cole a partir de qualquer célula: o conteúdo se espalha
        pelas colunas e cria as linhas que faltarem. Enter desce para a linha
        de baixo; a última linha em branco vira nova sozinha; ✕ remove a linha.</p>
        <input type="hidden" name="linhas" id="linhas">
        <button type="submit">▶ Criar usuários</button>
        <p class="mini">
          Senha padrão <code>Grupo@2026</code>, sem forçar troca no primeiro logon.<br>
          Login: <code>PRIMEIRO.ULTIMO</code>; se já existir, tenta
          <code>PRIMEIRO.PENULTIMO</code>, <code>PRIMEIRO.ANTEPENULTIMO</code> e depois
          <code>PRIMEIRO.ULTIMO2</code>, <code>3</code>…<br>
          Grupo: funções com GERENTE ou LIDER DE LOJA → <b>000013</b> (gerentes);
          as demais → <b>000012</b> (caixas do PDV).
        </p>
      </form>
    </div>
    """ + SCRIPT_GRADE
    return render(conteudo, aba="novo")


@app.route("/criar", methods=["POST"])
@exige_login
def criar():
    filial = (request.form.get("filial") or "").strip().upper()
    linhas, problemas = parse_linhas(request.form.get("linhas"))
    if not filial:
        return redirect(url_for("index", erro="Informe o código da filial."))
    if not re.fullmatch(r"[0-9A-Z]{6,20}", filial):
        return redirect(url_for("index", erro=(
            f"{filial!r} não parece um código de filial do Protheus "
            "(ex.: 01DOMA0001).")))
    if problemas:
        return redirect(url_for("index", erro=" | ".join(problemas[:5])))
    if not linhas:
        return redirect(url_for("index", erro="Nenhuma linha válida."))
    # Chrome fechado não bloqueia mais: o robô abre e loga sozinho (30/07/2026)
    if not _trava_execucao.acquire(blocking=False):
        return redirect(url_for("index", erro="Já existe uma execução em andamento — aguarde."))

    filial_nome = nome_da_filial(filial)
    try:
        with db() as con:
            cur = con.execute(
                "INSERT INTO execucoes (iniciada, filial, filial_nome) VALUES (?,?,?)",
                (agora(), filial, filial_nome))
            execucao_id = cur.lastrowid
            pendentes = []
            for u in linhas:
                repetido = con.execute(
                    "SELECT usuario FROM usuarios WHERE status='CRIADO' AND "
                    "replace(replace(replace(cpf,'.',''),'-',''),' ','')=?",
                    (so_digitos(u["cpf"]),)).fetchone()
                status = "PROCESSANDO"
                quando = None
                if repetido:
                    status = f"JÁ EXISTE (login {repetido['usuario']})"
                    quando = agora()
                cur = con.execute(
                    "INSERT INTO usuarios (execucao_id, filial, filial_nome, nome, cpf, "
                    "funcao, status, criado_em) VALUES (?,?,?,?,?,?,?,?)",
                    (execucao_id, filial, filial_nome, u["nome"], u["cpf"],
                     u["funcao"], status, quando))
                if not repetido:
                    pendentes.append({"id": cur.lastrowid, "filial": filial, **u})

        if pendentes:
            threading.Thread(target=rodar_execucao, args=(execucao_id, pendentes),
                             daemon=True).start()
        else:
            with db() as con:
                _finalizar(con, execucao_id, "CONCLUIDA")
            _trava_execucao.release()
    except Exception:
        _trava_execucao.release()
        raise

    return redirect(url_for("ver_execucao", execucao_id=execucao_id))


@app.route("/execucao/<int:execucao_id>")
@exige_login
def ver_execucao(execucao_id):
    with db() as con:
        ex = con.execute("SELECT * FROM execucoes WHERE id=?", (execucao_id,)).fetchone()
        usuarios = con.execute(
            "SELECT * FROM usuarios WHERE execucao_id=? ORDER BY id", (execucao_id,)).fetchall()
    if not ex:
        return redirect(url_for("historico"))
    rodando = ex["status"] == "RODANDO"
    situacao = ("⏳ Rodando... (a página se atualiza sozinha; cada usuário leva ~1 min)"
                if rodando else
                f"Resultado: {ex['criados']} criado(s), {ex['erros']} erro(s) de {ex['total']}")
    botao_parar = ""
    if rodando:
        botao_parar = f"""
        <form method="post" action="{url_for('parar_execucao', execucao_id=ex['id'])}"
              style="margin:0 0 12px 0"
              onsubmit="return confirm('Parar a execução nº {ex['id']} agora?\\n\\nQuem já foi criado continua criado no Protheus; quem ainda não foi fica com status de parado e pode ser recadastrado depois.')">
          <button type="submit" class="parar">■ Parar execução</button>
        </form>"""
    conteudo = f"""
    <div class="cartao">
      <h2>Execução nº {ex['id']} — filial {ex['filial']} {ex['filial_nome'] or ''}</h2>
      <p>Iniciada: {ex['iniciada']}
      {('| Finalizada: ' + ex['finalizada']) if ex['finalizada'] else ''}
      | Situação: <b>{ex['status']}</b></p>
      <p>{situacao}</p>
      {botao_parar}
      {render_tabela(usuarios)}
    </div>
    """
    return render(conteudo, aba="historico", refresh=rodando)


@app.route("/parar_execucao/<int:execucao_id>", methods=["POST"])
@exige_login
def parar_execucao(execucao_id):
    """
    Mata o processo do robô da execução em andamento (taskkill /T derruba
    também o chromedriver filho; o Chrome do Protheus NÃO é filho dele e
    continua aberto). Os usuários ainda PROCESSANDO ficam com status de
    parado; quem já foi criado permanece CRIADO. A tela do Protheus pode
    ficar com um formulário aberto — a próxima execução descarta sozinha
    (fechar_rotina/abandonar_formulario).
    """
    info = dict(_proc_atual)
    if info["execucao_id"] == execucao_id and info["proc"] is not None \
            and info["proc"].poll() is None:
        _paradas.add(execucao_id)
        subprocess.run(["taskkill", "/PID", str(info["proc"].pid), "/T", "/F"],
                       capture_output=True)
        # o robô morto deixa a trava de execução única para trás e bloquearia
        # a próxima execução por até 15 min — limpa aqui (nome da trava tem a
        # porta do Chrome; ver LOCK_PATH em robo/protheus_criar_usuario.py)
        try:
            os.remove(os.path.join(
                os.environ.get("TEMP", "."),
                f"protheus_autocadastro_{CHROME_DEBUG.rsplit(':', 1)[-1]}.lock"))
        except OSError:
            pass
    return redirect(url_for("ver_execucao", execucao_id=execucao_id))


@app.route("/historico")
@exige_login
def historico():
    with db() as con:
        execucoes = con.execute("SELECT * FROM execucoes ORDER BY id DESC LIMIT 300").fetchall()
    linhas = "".join(f"""
        <tr>
          <td><a href="{url_for('ver_execucao', execucao_id=e['id'])}">nº {e['id']}</a></td>
          <td>{e['iniciada']}</td><td>{e['finalizada'] or ''}</td>
          <td>{e['filial']} {e['filial_nome'] or ''}</td>
          <td>{e['total']}</td><td>{e['criados']}</td><td>{e['erros']}</td>
          <td class="st-{'CRIADO' if e['status'] == 'CONCLUIDA' else ('ERRO' if e['status'] in ('FALHOU', 'CANCELADO') else 'PROCESSANDO')}">{e['status']}</td>
        </tr>""" for e in execucoes)
    conteudo = f"""
    <div class="cartao">
      <h2>Histórico de execuções</h2>
      <div class="rolagem"><table>
        <tr><th>Execução</th><th>Iniciada</th><th>Finalizada</th><th>Filial</th>
        <th>Total</th><th>Criados</th><th>Erros</th><th>Situação</th></tr>
        {linhas or '<tr><td colspan="8">Nenhuma execução ainda.</td></tr>'}
      </table></div>
      <p class="mini">Clique no número da execução para ver todas as colunas de cada usuário.</p>
    </div>
    """
    return render(conteudo, aba="historico")


@app.route("/planilha")
@exige_login
def planilha():
    # esta tela mostra SÓ quem foi criado com sucesso (STATUS=CRIADO);
    # cadastros que falharam aparecem apenas no Histórico de execuções.
    filial_sel = request.args.get("filial", "")
    with db() as con:
        filiais = [r["filial"] for r in con.execute(
            "SELECT DISTINCT filial FROM usuarios WHERE status='CRIADO' ORDER BY filial")]
        if filial_sel:
            usuarios = con.execute(
                "SELECT * FROM usuarios WHERE status='CRIADO' AND filial=? ORDER BY nome",
                (filial_sel,)).fetchall()
        else:
            usuarios = con.execute(
                "SELECT * FROM usuarios WHERE status='CRIADO' ORDER BY filial, nome").fetchall()
    abas = f'<a href="{url_for("planilha")}" class="{"ativa" if not filial_sel else ""}">TODAS</a>'
    for f in filiais:
        cls = "ativa" if f == filial_sel else ""
        abas += f'<a href="{url_for("planilha", filial=f)}" class="{cls}">{f}</a>'
    aviso = request.args.get("aviso", "")
    conteudo = f"""
    {f'<div class="ok">{aviso}</div>' if aviso else ''}
    <div class="cartao">
      <h2>Usuários criados pela automação</h2>
      <p><a href="{url_for('exportar')}" class="botao">⬇ Exportar Excel</a></p>
      <div class="abas">{abas}</div>
      {render_tabela(usuarios, apagar=True)}
      <p class="mini">{len(usuarios)} usuário(s) criado(s)
      {'na filial ' + filial_sel if filial_sel else 'no total'}.
      Só aparecem aqui cadastros concluídos com sucesso — os que falharam
      ficam no Histórico de execuções.<br>
      O ✕ apaga a linha <b>somente do banco do site</b> — o usuário continua
      existindo no Protheus; o CPF dele fica liberado para um novo cadastro.</p>
    </div>
    """
    return render(conteudo, aba="planilha")


@app.route("/apagar_usuario/<int:usuario_id>", methods=["POST"])
@exige_login
def apagar_usuario(usuario_id):
    """Apaga o registro SÓ do banco do site (o Protheus não é tocado).
    Com isso a trava de CPF repetido deixa de valer para essa pessoa."""
    with db() as con:
        u = con.execute("SELECT * FROM usuarios WHERE id=?", (usuario_id,)).fetchone()
        if not u:
            return redirect(url_for("planilha"))
        con.execute("DELETE FROM usuarios WHERE id=?", (usuario_id,))
    aviso = (f"{u['nome']} ({u['usuario'] or 'sem login'}) apagado do banco do "
             "site. O Protheus não foi alterado.")
    return redirect(url_for("planilha", filial=request.args.get("filial", ""),
                            aviso=aviso))


@app.route("/exportar")
@exige_login
def exportar():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cab = ["NOME", "CPF", "FUNCAO", "USUARIO", "SENHA", "CONFIRME_SENHA",
           "GRUPO_CODIGO", "GRUPO_NOME", "PRIORIZA_GRUPO", "FORCAR_TROCA_SENHA",
           "CODIGO_BANCO", "ID_USUARIO", "STATUS", "QUANDO"]
    with db() as con:
        # exporta o mesmo que a tela "Usuários criados pela automação":
        # só cadastros concluídos com sucesso (STATUS=CRIADO)
        filiais = [r["filial"] for r in con.execute(
            "SELECT DISTINCT filial FROM usuarios WHERE status='CRIADO' "
            "ORDER BY filial")] or ["(vazio)"]
        for f in filiais:
            ws = wb.create_sheet(title=(f[:31] or "(vazio)"))
            ws.append(cab)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="DDEBF7")
            for u in con.execute("SELECT * FROM usuarios WHERE status='CRIADO' "
                                 "AND filial=? ORDER BY nome", (f,)):
                ws.append([u["nome"], u["cpf"], u["funcao"], u["usuario"], u["senha"],
                           u["senha"], u["grupo_codigo"], u["grupo_nome"],
                           u["prioriza_grupo"], u["forcar_troca_senha"],
                           u["codigo_banco"], u["id_usuario"], u["status"],
                           u["criado_em"] or ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"usuarios_criados_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ----------------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    with db() as con:
        con.execute("UPDATE usuarios SET status='ERRO: site reiniciado no meio' "
                    "WHERE status='PROCESSANDO'")
        con.execute("UPDATE execucoes SET status='FALHOU', finalizada=? WHERE status='RODANDO'",
                    (agora(),))
    print(f"AutoCadastro no ar: http://0.0.0.0:{PORTA}")
    print(f"Login do site: {SITE_USUARIO} / {'(senha configurada)' if SITE_SENHA else '(SEM SENHA)'}")
    print(f"Chrome do Protheus em {CHROME_DEBUG}: "
          f"{'CONECTADO' if chrome_conectado() else 'NAO CONECTADO'}")
    try:
        from waitress import serve
        serve(app, host="0.0.0.0", port=PORTA, threads=8)
    except ImportError:
        app.run(host="0.0.0.0", port=PORTA)
